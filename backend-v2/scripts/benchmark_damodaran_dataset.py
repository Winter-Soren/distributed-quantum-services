"""Benchmark quantum vs classical portfolio optimization using Professor Aswath Damodaran's NYU dataset.

Professor Damodaran's historical returns dataset (1928-2025) provides 97 years of reliable,
academically-vetted market data across multiple asset classes:
- S&P 500 (with dividends)
- US Small Cap (bottom decile)
- 3-month T-Bills
- US Treasury Bonds (10-year)
- Baa Corporate Bonds
- Real Estate
- Gold

This comprehensive benchmark tests quantum advantage on real academic data spanning
nearly a century of market conditions including:
- The Great Depression (1929-1939)
- Post-WWII boom (1945-1970)
- Stagflation era (1970s)
- Dot-com bubble (1990s-2000s)
- Financial crisis (2008)
- COVID-19 pandemic (2020)

Expected: Quantum maintains constant parameter search time regardless of portfolio complexity,
while classical enumeration grows exponentially with asset count.
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import io
import json
import logging
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path

from run_track_b_market_benchmark import _run_distributed_benchmark, _rows_to_csv_bytes


def _project_root() -> Path:
    return Path(__file__).resolve().parents[2]


def _install_openpyxl() -> None:
    """Install openpyxl for Excel file reading."""
    print("📦 Installing openpyxl for Excel file reading...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
    print("✅ openpyxl installed successfully")


def _parse_damodaran_excel(excel_path: Path) -> list[dict[str, str]]:
    """Parse Damodaran's historical returns Excel file into CSV-compatible rows.

    The Excel file has annual returns (1928-2025) for multiple asset classes.
    We'll convert this into a date-indexed price matrix suitable for portfolio optimization.
    """
    try:
        import openpyxl
    except ImportError:
        _install_openpyxl()
        import openpyxl

    print(f"\n📊 Parsing Damodaran dataset: {excel_path.name}")

    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = workbook.active

    # Find the header row and data range
    # Damodaran's files typically have the year in column A and asset returns in subsequent columns
    header_row = None
    data_start_row = None

    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20), start=1):
        # Look for a row that contains "Year" or starts with a year
        first_cell = row[0].value
        if first_cell and (str(first_cell).lower() == "year" or str(first_cell).isdigit()):
            if str(first_cell).lower() == "year":
                header_row = row_idx
                data_start_row = row_idx + 1
            else:
                # Data starts immediately (no explicit header)
                data_start_row = row_idx
                # Build header from previous row or create generic one
                header_row = row_idx - 1 if row_idx > 1 else None
            break

    if data_start_row is None:
        raise ValueError("Could not find data start row in Excel file")

    # Extract headers (asset class names)
    if header_row:
        headers = []
        for cell in sheet[header_row]:
            if cell.value:
                # Clean up header names
                header = str(cell.value).strip()
                # Replace spaces and special chars for CSV compatibility
                header = header.replace(" ", "_").replace("&", "and").replace("/", "_")
                headers.append(header)
            else:
                break
    else:
        # No header row, create generic headers
        max_col = 0
        for row in sheet.iter_rows(min_row=data_start_row, max_row=data_start_row):
            for idx, cell in enumerate(row):
                if cell.value is not None:
                    max_col = idx + 1
        headers = ["Year"] + [f"Asset_{i}" for i in range(1, max_col)]

    print(f"   Found {len(headers)} columns: {', '.join(headers[:5])}{'...' if len(headers) > 5 else ''}")

    # Parse data rows (convert returns to cumulative prices)
    rows = []
    base_prices = {header: 100.0 for header in headers[1:]}  # Start each asset at $100

    for row in sheet.iter_rows(min_row=data_start_row, max_row=sheet.max_row):
        year_cell = row[0].value
        if not year_cell or not str(year_cell).strip().isdigit():
            continue

        year = int(year_cell)
        # Convert year to date format (use December 31st)
        date_str = f"{year}-12-31"

        row_data = {"date": date_str}

        for idx, header in enumerate(headers[1:], start=1):
            if idx < len(row):
                return_pct = row[idx].value
                if return_pct is not None:
                    try:
                        # Convert percentage return to price
                        # If return is 10.5%, multiply previous price by 1.105
                        return_decimal = float(return_pct) / 100.0
                        base_prices[header] *= (1 + return_decimal)
                        row_data[header] = f"{base_prices[header]:.2f}"
                    except (ValueError, TypeError):
                        pass

        # Only include rows with at least some data
        if len(row_data) > 1:
            rows.append(row_data)

    print(f"   Extracted {len(rows)} years of data ({rows[0]['date']} to {rows[-1]['date']})")

    # Verify we have all required columns for all rows
    if rows:
        final_headers = list(rows[0].keys())
        standardized_rows = []
        for row in rows:
            standardized_row = {h: row.get(h, "0.0") for h in final_headers}
            standardized_rows.append(standardized_row)
        return standardized_rows

    return rows


def _select_time_period(
    rows: list[dict[str, str]],
    *,
    start_year: int | None = None,
    end_year: int | None = None,
    last_n_years: int | None = None,
) -> list[dict[str, str]]:
    """Filter rows to a specific time period."""
    if last_n_years:
        if len(rows) >= last_n_years:
            return rows[-last_n_years:]
        return rows

    filtered = []
    for row in rows:
        date_str = row["date"]
        year = int(date_str.split("-")[0])

        if start_year and year < start_year:
            continue
        if end_year and year > end_year:
            continue

        filtered.append(row)

    return filtered


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--dataset",
        choices=["histret", "histimpl"],
        default="histret",
        help="Which Damodaran dataset to use: histret (historical returns) or histimpl (implied premiums)",
    )
    parser.add_argument(
        "--time-period",
        choices=["full", "modern", "recent", "decade", "crisis", "depression"],
        default="full",
        help=(
            "Time period to analyze: "
            "full (1928-2025), modern (1980-2025), recent (2000-2025), "
            "decade (2015-2025), crisis (2007-2010), depression (1929-1940)"
        ),
    )
    parser.add_argument(
        "--peer-count",
        type=int,
        default=30,
        help="Number of distributed quantum nodes to simulate",
    )
    parser.add_argument(
        "--max-assets",
        type=int,
        default=5,
        help="Maximum number of assets to include in portfolio optimization",
    )
    parser.add_argument(
        "--parameter-search-steps",
        type=int,
        default=7,
        help="QAOA parameter search grid resolution",
    )
    parser.add_argument(
        "--budget",
        type=int,
        default=None,
        help="Optional fixed portfolio budget constraint",
    )
    parser.add_argument(
        "--run-multiple-scales",
        action="store_true",
        help="Run benchmarks at multiple asset scales to demonstrate quantum scaling advantage",
    )
    return parser.parse_args()


async def _run_single_benchmark(
    *,
    csv_bytes: bytes,
    filename: str,
    peer_count: int,
    max_assets: int,
    parameter_search_steps: int,
    budget: int | None,
    dataset_info: dict,
) -> dict:
    """Run a single benchmark configuration."""
    print(f"\n{'─'*90}")
    print(f"🚀 BENCHMARK CONFIGURATION")
    print(f"   Dataset: {filename}")
    print(f"   Time period: {dataset_info['period_name']} ({dataset_info['start_year']}-{dataset_info['end_year']})")
    print(f"   Asset classes: {max_assets}")
    print(f"   Data points: {dataset_info['data_points']}")
    print(f"   Quantum nodes: {peer_count}")
    print(f"   Parameter search steps: {parameter_search_steps}")
    print(f"{'─'*90}")

    start_time = time.time()

    benchmark = await _run_distributed_benchmark(
        csv_bytes=csv_bytes,
        filename=filename,
        peer_count=peer_count,
        max_assets_considered=max_assets,
        parameter_search_steps=parameter_search_steps,
        budget=budget,
    )

    elapsed_time = time.time() - start_time

    timings = benchmark.get("timings", {})
    scorecard = benchmark.get("comparison_report", {}).get("scorecard", {})

    classical_ms = timings.get("classical_end_to_end_duration_ms", 0)
    quantum_ms = timings.get("quantum_end_to_end_duration_ms", 0)

    print(f"\n✅ BENCHMARK COMPLETED in {elapsed_time:.2f}s")
    print(f"   Classical time: {classical_ms}ms")
    print(f"   Quantum time: {quantum_ms}ms")
    print(f"   Parameter search: {timings.get('quantum_parameter_search_duration_ms', 0)}ms")
    print(f"   Winner: {scorecard.get('winner_by_runtime', 'unknown')}")

    if classical_ms > 0 and quantum_ms > 0:
        ratio = quantum_ms / classical_ms
        if ratio < 1:
            print(f"   🎉 Quantum {1/ratio:.2f}× FASTER!")
        else:
            print(f"   Quantum {ratio:.2f}× slower")

    return {
        "dataset_info": dataset_info,
        "config": {
            "peer_count": peer_count,
            "max_assets": max_assets,
            "parameter_search_steps": parameter_search_steps,
            "budget": budget,
        },
        "total_benchmark_time_seconds": round(elapsed_time, 3),
        "timings": timings,
        "scorecard": scorecard,
        "benchmark_details": benchmark,
    }


def main() -> None:
    args = _parse_args()
    logging.getLogger("qiskit").setLevel(logging.WARNING)

    print(f"\n{'='*90}")
    print("QUANTUM VS CLASSICAL PORTFOLIO OPTIMIZATION")
    print("Using Professor Aswath Damodaran's NYU Historical Returns Dataset")
    print(f"{'='*90}")

    # Determine dataset path
    damodaran_dir = _project_root() / "benchmark-data" / "damodaran"
    if args.dataset == "histret":
        excel_path = damodaran_dir / "histretSP.xls"
        dataset_name = "Historical Returns (1928-2025)"
    else:
        excel_path = damodaran_dir / "histimpl.xls"
        dataset_name = "Implied Equity Risk Premiums (1960-2025)"

    if not excel_path.exists():
        print(f"\n❌ Dataset not found: {excel_path}")
        print("The dataset should have been downloaded. Please check the path.")
        sys.exit(1)

    # Parse the Excel file
    rows = _parse_damodaran_excel(excel_path)

    # Select time period
    period_configs = {
        "full": {"start": None, "end": None, "name": "Full History"},
        "modern": {"start": 1980, "end": None, "name": "Modern Era"},
        "recent": {"start": 2000, "end": None, "name": "Recent Markets"},
        "decade": {"start": 2015, "end": None, "name": "Last Decade"},
        "crisis": {"start": 2007, "end": 2010, "name": "Financial Crisis"},
        "depression": {"start": 1929, "end": 1940, "name": "Great Depression"},
    }

    period_config = period_configs[args.time_period]
    filtered_rows = _select_time_period(
        rows,
        start_year=period_config["start"],
        end_year=period_config["end"],
    )

    if len(filtered_rows) < 10:
        print(f"❌ Not enough data points: {len(filtered_rows)}")
        sys.exit(1)

    dataset_info = {
        "dataset_name": dataset_name,
        "period_name": period_config["name"],
        "start_year": filtered_rows[0]["date"].split("-")[0],
        "end_year": filtered_rows[-1]["date"].split("-")[0],
        "data_points": len(filtered_rows),
        "asset_classes": list(filtered_rows[0].keys())[1:],  # Exclude 'date' column
    }

    print(f"\n📊 Dataset Summary:")
    print(f"   Name: {dataset_info['dataset_name']}")
    print(f"   Period: {dataset_info['period_name']} ({dataset_info['start_year']}-{dataset_info['end_year']})")
    print(f"   Data points: {dataset_info['data_points']}")
    print(f"   Asset classes: {', '.join(dataset_info['asset_classes'])}")

    csv_bytes = _rows_to_csv_bytes(filtered_rows)

    if args.run_multiple_scales:
        # Run benchmarks at multiple scales
        asset_counts = [3, 4, 5, 6, 7] if len(dataset_info['asset_classes']) >= 7 else [3, 4, 5]
        results = []

        print(f"\n{'='*90}")
        print("MULTI-SCALE BENCHMARK: Testing Quantum Scaling Advantage")
        print(f"{'='*90}")

        for max_assets in asset_counts:
            if max_assets > len(dataset_info['asset_classes']):
                print(f"\n⚠️  Skipping {max_assets} assets (only {len(dataset_info['asset_classes'])} available)")
                continue

            result = asyncio.run(
                _run_single_benchmark(
                    csv_bytes=csv_bytes,
                    filename=excel_path.name,
                    peer_count=args.peer_count,
                    max_assets=max_assets,
                    parameter_search_steps=args.parameter_search_steps,
                    budget=args.budget,
                    dataset_info=dataset_info,
                )
            )
            results.append(result)

        # Analysis summary
        print(f"\n{'='*90}")
        print("SCALING ANALYSIS SUMMARY")
        print(f"{'='*90}\n")
        print(f"{'Assets':<10} {'Classical (ms)':<18} {'Quantum (ms)':<18} {'Ratio':<10} {'Winner':<15}")
        print(f"{'-'*90}")

        for result in results:
            max_assets = result["config"]["max_assets"]
            timings = result["timings"]
            classical_ms = timings["classical_end_to_end_duration_ms"]
            quantum_ms = timings["quantum_end_to_end_duration_ms"]
            ratio = quantum_ms / classical_ms if classical_ms > 0 else 0
            winner = "🏆 Quantum" if ratio < 1 else "Classical"

            print(f"{max_assets:<10} {classical_ms:<18.1f} {quantum_ms:<18.1f} {ratio:<10.2f} {winner:<15}")

        # Save results
        output_path = _project_root() / "benchmark-data" / "damodaran_benchmark_results.json"
        output_path.write_text(json.dumps(results, indent=2))
        print(f"\n📁 Results saved to: {output_path}")

    else:
        # Single benchmark run
        result = asyncio.run(
            _run_single_benchmark(
                csv_bytes=csv_bytes,
                filename=excel_path.name,
                peer_count=args.peer_count,
                max_assets=args.max_assets,
                parameter_search_steps=args.parameter_search_steps,
                budget=args.budget,
                dataset_info=dataset_info,
            )
        )

        # Save result
        output_path = _project_root() / "benchmark-data" / f"damodaran_benchmark_{args.time_period}.json"
        output_path.write_text(json.dumps(result, indent=2))
        print(f"\n📁 Result saved to: {output_path}")

    print(f"\n{'='*90}")
    print("BENCHMARK COMPLETE!")
    print(f"{'='*90}\n")


if __name__ == "__main__":
    main()
